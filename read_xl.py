
from openpyxl import load_workbook as lw
from datetime import datetime
try: 
    from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string
# End of Importing


wb = lw('Attendance.xlsx', data_only= True)	#Reading the Workbook
sheets = wb.sheetnames 					#Getting sheet names
w_sh = wb[sheets[0]]					#Reading the First sheet

fh = open('./store.txt')

collection = list()
i = 2
while w_sh.cell(row=i,column=2).value != None:
	collection.append( (str(w_sh.cell(row=i,column=2).value.strip()), str(w_sh.cell(row=i,column=6).value.strip())) )
	i = i + 1
#Collection of students constructed


#Checking Attended Pupils
attended = list()

for line in fh:
	if line == None: continue
	line = line.strip()

	A = line.split('-')		
	for i in collection:	#Query by Roll number
		roll_ = i[1].split('-')
		if A[-1] == roll_[-1]:
			attended.append(i)
	
	for i in collection:	#Query by Full Name
		name_ = i[0].upper()
		if name_ == line.upper():
			attended.append(i)		
fh.close()

#Function to Remove duplicate from attendance list
def removeDuplicates(lst): 
      return [t for t in (set(tuple(i) for i in lst))] 


w_sh = wb[sheets[1]]	#Writing attendance in the Second Sheet

i = 5


while w_sh.cell(row=2, column=i).value != None:	#while date field is empty

	i = i + 1		#Increament
	
	verify = w_sh.cell(row=2, column=i).value			#Grabbing New date if available
	current_date = datetime.today().strftime('%d-%m-%Y')	#Grabbing Current date from time library
	
	if  verify == None and verify != current_date:	#If today's attendance is not written
		col_letter = get_column_letter(i)
		index = col_letter + '2'
		#print(index)
		w_sh[index] = current_date
		wb.save('Attendance.xlsx')
		break
		
	elif verify == current_date:
		col_letter = get_column_letter(i)
		index = col_letter + '2'
		#print(index)
		#print('Attendance is Up to Date')			#Else
		break


col = index[:1]
co = column_index_from_string(col)
row = index[1:5]
ro = int(row) + 1


while w_sh.cell(row=ro, column = co).value != 'Test':
	
	roll = w_sh.cell(row=ro, column=2).value	
	name_ = w_sh.cell(row=ro, column=4).value
	name = name_.strip()
	index = col + str(ro)
	
	jora = (name, roll)
	#print(jora)
	if jora in  removeDuplicates(attended):
		w_sh[index] = 'P'
	
	else:
		w_sh[index] = ''
		
		
	ro = ro + 1

print(attended)
print('Attendance is now Up to Date')
wb.save('Attendance.xlsx')
	
